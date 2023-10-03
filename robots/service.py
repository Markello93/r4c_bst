from django.db.models import Count
from django.utils import timezone
from openpyxl.workbook import Workbook

from robots.models import Robot


def from_db_to_excel():
    """Создает excel файл и фильтрует данные о роботах за последнюю неделю."""
    week = timezone.now() - timezone.timedelta(days=7)
    wb = Workbook()
    model_data = (
        Robot.objects.filter(created__gte=week)
            .values('model', 'version')
            .annotate(count=Count('id'))
            .distinct()
    )

    for item in model_data:
        sheet = (
            wb.get_sheet_by_name(item['model'])
            if item['model'] in wb.sheetnames
            else wb.create_sheet(item['model'])
        )
        if 'A1' not in sheet:
            sheet['A1'] = 'Модель'
            sheet['B1'] = 'Версия'
            sheet['C1'] = 'Количество за неделю'

        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=item['model'])
        sheet.cell(row=row, column=2, value=item['version'])
        sheet.cell(row=row, column=3, value=item['count'])

    wb.remove(wb['Sheet'])
    return wb, model_data
